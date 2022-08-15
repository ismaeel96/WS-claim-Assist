#selenium, pandas, lxml, time, html5lib
import time
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
from selenium.webdriver.chrome.service import Service

from selenium.webdriver.chrome.options import Options

#opens chrome
#driver = webdriver.Chrome('./chromedriver')
#driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

#option to keep browser open
#chrome_options = Options()
#chrome_options.add_experimental_option("detach", True)

Excel_File_Path = 'VinList.xlsx'
path = Path(Excel_File_Path)

Start_Column = 1
Check_Start_Row = 2

#vinList = ['1C4SJVFJ1NS103023','1C4SJVFJ1NS103023','1C4SJVDT6NS110263','1C4SJVDT3NS102718','1C4SJVDT9NS104439','1C4SJVDT7NS104178','1C4SJVFJ9NS101827','1C4SJVFJ7NS108484','1C4SJVDT6NS104026','1C4SJVDT6NS131985','1C4SJVDT9NS103761','1C4SJVGJ3NS101739','1C4SJVGJ3NS101739','1C4SJVDT3NS106672','1C4SJVDT3NS103982','1C4SJVDT6NS101434','1C4SJVDT6NS101434','1C4SJVDT6NS101434','1C4SJVDT6NS101434','1C4SJVFJ0NS107130','1C4SJVDT4NS101691','1C4SJVDT2NS105660','1C4SJVDT2NS105660','1C4SJVGJ5NS101533','1C4SJVGJ3NS115625','1C4SJVDT3NS104372','1C4SJVDT2NS104931','1C4SJVDT2NS104931','1C4SJVDT6NS104088','1C4SJVDT1NS104886','1C4SJVDT7NS105279','1C4SJVDT8NS103122','1C4SJVDT8NS103122','1C4SJVDT9NS103033','1C4SJVDT8NS107512','1C4SJVGJ5NS110197','1C4SJVDT7NS106612','1C4SJVGJ9NS101227','1C4SJVDT1NS103673']
#claimList=['039840','039840','657362','758780','636675','183029','537066','748895','342780','182977','16774C','318069','318069','034132','001369','071988','071988','071988','071988','768720','265941','C20078','D00780','013806','098504','228829','441425','A40605','354140','660492','574547','142479','142479','244156','597720','A51246','433961','A88588','W81202']
vinList=[]
claimList=[]

if path.is_file():
    print(f'The file {Excel_File_Path} exists')
    workbook = load_workbook(filename="VinList.xlsx")
else:
    print(f'The file {Excel_File_Path} does not exist')
    workbook = Workbook()
    print(f'Created file:{Excel_File_Path}')
sheet = workbook.active

username = str(sheet.cell(row = 1, column = 1).value)
password = str(sheet.cell(row = 1, column = 2).value)
loopIndex=0

while sheet.cell(row = Check_Start_Row, column = 1).value != None:
  #print(sheet.cell(row = 2, column = Start_Column).value)
  #print("column: ", Start_Column)
  vinList.append(str(sheet.cell(row = Check_Start_Row, column = 1).value))
  claimList.append(str(sheet.cell(row=Check_Start_Row, column=2).value))
  Check_Start_Row += 1

def button_clicked(browser):
    global loopIndex
    if (browser.find_element(By.ID, 'next_claim_button').get_attribute('value')=="true"):
        loopIndex+=1
        if (loopIndex > (len(vinList)-1)):
            loopIndex=0
            browser.execute_script("alert('already at last claim, Going to first claim in 3 seconds');")
            time.sleep(3)
            search_for_claim(browser)
        return True
    if (browser.find_element(By.ID, 'previous_claim_button').get_attribute('value')=="true"):
        loopIndex-=1
        if (loopIndex < 0):
            loopIndex=(len(vinList)-1)
            browser.execute_script("alert('Already at first claim, Going to last claim in 3 seconds');")
            time.sleep(3)
            search_for_claim(browser)
        return True

print("claimTabView('{vinFirst}','{vinLast}','','{claimNum}','1','C','','');".format( vinFirst=vinList[0][0:9:1] , vinLast=vinList[0][9:17:1] , claimNum=claimList[0] ))
def Open_GCS():
    try:
        # option to keep browser open
        chrome_options = Options()
        #chrome_options = webdriver.ChromeOptions()
        chrome_options.binary_location = "drivers\chrome-win\chrome.exe"
        #service_obj = Service("/drivers/chromedriver.exe")
        driverPath = "drivers\chromedriver.exe"

        #chrome_options.add_experimental_option("detach", True)

        #chrome_options = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        ##chrome_options.add_argument("user-data-dir=C:\\Users\\Desktop\\AppData\\Local\\Google\\Chrome\\User Data")
        global browser # this will prevent the browser variable from being garbage collected

        #this adds user optiosn for chrome
        #browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        browser = webdriver.Chrome(executable_path=driverPath,options=chrome_options)
        # Connect
        browser.set_window_size(900, 900)

        browser.get("https://sts.fiatgroup.com/adfs/ls/IdpInitiatedSignOn.aspx?RelayState=RPID%3Dhttps%253A%252F%252Fwww.esupplierconnect.com%252Fsaml2%252Fsp%252Facs%26RelayState%3Dfiat&fhr=default")
        #browser.get("https://webprod.extra.chrysler.com/VehiInquiryWeb/viTabsExecute?urlFlag=")


        #login
        try:
            element = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.ID, "userNameInput"))
            )
        finally:
            browser.find_element(By.ID, 'userNameInput').send_keys(username, Keys.TAB, password, Keys.ENTER)
            #browser.find_element(By.Id("showMoreappBtnTxt")).Click()
        time.sleep(15)
        start_search_for_claim(browser)

        #browser.get("https://webprod.extra.chrysler.com/VehiInquiryWeb/viTabsExecute?urlFlag=")
        #browser.find_element(By.Id("showMoreappBtnTxt")).Click()
        #https://webprod.extra.chrysler.com/VehiInquiryWeb/searchExecute
        #browser.find_element(By.NAME, 'username').send_keys('MYEMAIL', Keys.TAB, 'MYPW', Keys.ENTER)
    except Exception as e:
        print (e, 'GCS')

def start_search_for_claim(browser):
    browser.execute_script("showMoreApps();")
    browser.find_element_by_xpath('//span[contains(text(),"Supplier Warranty Management - WIS,EWT,GCS,QNA")]').click()
    time.sleep(5)
    browser.get("https://webprod.extra.chrysler.com/VehiInquiryWeb/viTabsExecute?urlFlag=")
    time.sleep(5)

    search_for_claim(browser)

    #browser.find_element_by_link_text('Supplier Warranty Management - WIS,EWT,GCS,QNA').click()
    #browser.find_element_by_link_text('System Links').click()
    #browser.find_element_by_link_text('GCS - Vehicle Claims Inquiry').click()
    #browser.find_element(By.ID, 'viVin2').send_keys(vinList[0], Keys.ENTER)
  ##  JSCommand="claimTabView('{vinFirst}','{vinLast}','','{claimNum}','1','C','','');".format( vinFirst=vinList[0][0:9:1] , vinLast=vinList[0][9:17:1] , claimNum=claimList[0])
  ##  browser.execute_script(str(JSCommand))
  ##  browser.execute_script("loadTabAction(6);")
 ##   time.sleep(1)
##    mycode = """document.getElementById('tabMain').insertAdjacentHTML('afterend', '<div style="position: fixed; top: 0; right: 0; z-index:999; float: right;"><button type="button" style="width: 125px; height: 100px;"id="previous_claim_button" value="false" onclick="document.getElementById(\\'previous_claim_button\\').value = \\'true\\';">Previous Claim</button><button type="button" style="width: 125px; height: 100px;" id="next_claim_button"  value="false" onclick="document.getElementById(\\'next_claim_button\\').value = \\'true\\';">Next Claim</button></div>');"""
##    browser.execute_script(str(mycode))

def search_for_claim(browser):
    while True:
        JSCommand = "claimTabView('{vinFirst}','{vinLast}','','{claimNum}','1','C','','');".format(
            vinFirst=vinList[loopIndex][0:9:1], vinLast=vinList[loopIndex][9:17:1], claimNum=claimList[loopIndex])
        browser.execute_script(str(JSCommand))
        browser.execute_script("loadTabAction(6);")
        time.sleep(.5)

        mycode = """document.getElementById('tabMain').insertAdjacentHTML('afterend', '<div style="position: fixed; top: 0; right: 0; z-index:999; float: right;"><button type="button" style="width: 125px; height: 100px;"id="previous_claim_button" value="false" onclick="document.getElementById(\\'previous_claim_button\\').value = \\'true\\';">Previous Claim</button><button type="button" style="width: 125px; height: 100px;" id="next_claim_button"  value="false" onclick="document.getElementById(\\'next_claim_button\\').value = \\'true\\';">Next Claim</button><h1>1/85</h1></div>');"""
        browser.execute_script(str(mycode))
        browser.execute_script("document.querySelector('#gview_clmnarrGrid > div.ui-jqgrid-bdiv').style = 'height:500px;width:750px;'")
        WebDriverWait(browser, timeout=100000).until(button_clicked)



## double pound is working code
##    htmlSRC = browser.page_source
    #soup = BeautifulSoup(htmlSRC, 'html.parser')

    #div = soup.select_one("id#clmnarrGrid")
    #print(div)
    #table = pd.read_html(div)

    #HTMLTableRead = pd.read_html(browser.page_source, attrs={'id': 'clmnarrGrid'})
  ##  pd.set_option('display.max_colwidth', None)
  ##  pd.set_option("max_columns", None)  # show all cols
  ##  pd.set_option('max_colwidth', None)  # show full width of showing cols

  ##  HTMLTableRead = pd.read_html(htmlSRC, attrs={'id': 'clmnarrGrid'})
  ##  print(HTMLTableRead)

  ##  dataFromTables = HTMLTableRead[0]
 ##   print(dataFromTables)



Open_GCS()



#opens website
#driver.get("https://webprod.extra.chrysler.com/VehiInquiryWeb/viTabsExecute?urlFlag=")



#print(driver.title)


# Press the green button in the gutter to run the script.
#if __name__ == '__main__':


# See PyCharm help at https://www.jetbrains.com/help/pycharm/


#<div style="position: fixed; top: 0; right: 0; z-index:999; float: right;">
 #   <button type="button" style="width: 125px; height: 100px;" id="previous_claim_button" value="false" onclick="document.getElementById('previous_claim_button').value = 'true';">Previous Claim</button>
  #  <button type="button" style="width: 125px; height: 100px;" id="next_claim_button"  value="false" onclick="document.getElementById('next_claim_button').value = 'true';">Next Claim</button>
#</div>
#document.getElementById('tabMain').insertAdjacentHTML('afterend', '<div style="position: fixed; top: 0; right: 0; z-index:999; float: right;"><button type="button" style="width: 125px; height: 100px;"id="previous_claim_button" value="false" onclick="document.getElementById(\'previous_claim_button\').value = \'true\';">Previous Claim</button><button type="button" style="width: 125px; height: 100px;" id="next_claim_button"  value="false" onclick="document.getElementById(\'next_claim_button\').value = \'true\';">Next Claim</button></div>');
